import re
import PyPDF2
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os

# Function to extract text from PDF
def extract_pdf_text(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
    return text

# Function to extract required details using regex
def extract_details(text):
    details = {}

    try:
        # Extract company name (assumes the first uppercase block near the top is the company name)
        company_name_match = re.search(r'^[A-Z][A-Z &,-\.]+(?:\s+Ltd.*)?(?=\s+At\.)', text, re.MULTILINE)
        details['Company Name'] = company_name_match.group(0).strip() if company_name_match else ""

        # Extracting other fields
        details['Invoice No'] = re.search(r'Invoice No\.\s*[:\-]?\s*(\d+)', text).group(1) if re.search(r'Invoice No\.\s*[:\-]?\s*(\d+)', text) else ""
        details['Date of Invoice'] = re.search(r'Date of Invoice\s*[:\-]?\s*([\d-]+)', text).group(1) if re.search(r'Date of Invoice\s*[:\-]?\s*([\d-]+)', text) else ""
        details['GSTIN NO'] = re.search(r'GSTIN\s*NO\s*[:\-]?\s*([\w\d]+)', text).group(1) if re.search(r'GSTIN\s*NO\s*[:\-]?\s*([\w\d]+)', text) else ""
        details['GSTIN'] = re.search(r'GSTIN\s*[:\-]?\s*([\w\d]+)', text).group(1) if re.search(r'GSTIN\s*[:\-]?\s*([\w\d]+)', text) else ""
        details['HSN/SAC code'] = re.search(r'HSN/SAC\s*[:\-]?\s*(\d+)', text).group(1) if re.search(r'HSN/SAC\s*[:\-]?\s*(\d+)', text) else ""
        
        # Shipped to: Improved regex for multiline addresses
        shipped_to_match = re.search(r'Shipped to\s*[:\-]?\s*([\s\S]+?)\s*FSSAI', text)
        details['Shipped to'] = shipped_to_match.group(1).strip() if shipped_to_match else ""

        # Goods Description
        goods_description_match = re.search(r'DESCRIPTION OF GOODS\s*[:\-]?\s*([\s\S]+?)\s*HSN/SAC', text)
        details['Goods Description'] = goods_description_match.group(1).strip() if goods_description_match else ""

        # Bags, Quintal, Rate, Amount
        details['Bags'] = re.search(r'BAGS\s*[:\-]?\s*(\d+)', text).group(1) if re.search(r'BAGS\s*[:\-]?\s*(\d+)', text) else ""
        details['Quintal'] = re.search(r'QUINTAL\s*[:\-]?\s*([\d.]+)', text).group(1) if re.search(r'QUINTAL\s*[:\-]?\s*([\d.]+)', text) else ""
        details['Rate'] = re.search(r'RATE\s*[:\-]?\s*([\d.,]+)', text).group(1) if re.search(r'RATE\s*[:\-]?\s*([\d.,]+)', text) else ""
        details['Amount'] = re.search(r'AMOUNT\(`\s*\)\s*[:\-]?\s*([\d.,]+)', text).group(1) if re.search(r'AMOUNT\(`\s*\)\s*[:\-]?\s*([\d.,]+)', text) else ""

        # Transport, Vehicle No, Licence No, Mobile No
        details['Transport'] = re.search(r'Transport\s*[:\-]?\s*([\s\S]+?)\s+Despatch Date', text).group(1).strip() if re.search(r'Transport\s*[:\-]?\s*([\s\S]+?)\s+Despatch Date', text) else ""
        details['Vehicle No'] = re.search(r'Vehicle No\.\s*[:\-]?\s*([\w\d]+)', text).group(1) if re.search(r'Vehicle No\.\s*[:\-]?\s*([\w\d]+)', text) else ""
        details['Licence No'] = re.search(r'Licence No\s*[:\-]?\s*([\w\d]+)', text).group(1) if re.search(r'Licence No\s*[:\-]?\s*([\w\d]+)', text) else ""
        details['Mobile No'] = re.search(r'Mobile No\s*[:\-]?\s*([\d]+)', text).group(1) if re.search(r'Mobile No\s*[:\-]?\s*([\d]+)', text) else ""

    except Exception as e:
        print(f"Error extracting data: {e}")
    
    return details

# Function to write details to Excel (appending if file exists)
def write_to_excel(details, excel_path):
    # Check if the Excel file exists
    file_exists = os.path.exists(excel_path)

    # Create or load the workbook
    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        # Write header
        headers = list(details.keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    # Write data
    ws.append(list(details.values()))

    wb.save(excel_path)
    print(f"Data successfully written to {excel_path}")

# Main execution
def main():
    print("Please select the PDF file:")
    Tk().withdraw()  # Hides the root window
    pdf_path = askopenfilename(filetypes=[("PDF files", "*.pdf")])

    if pdf_path:
        text = extract_pdf_text(pdf_path)
        details = extract_details(text)
        excel_path = pdf_path.replace('.pdf', 'details.xlsx')
        write_to_excel(details, excel_path)
    else:
        print("No file selected.")

if __name__ == "__main__":
    main()
